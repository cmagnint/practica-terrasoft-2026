import os
import sys
from datetime import datetime

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

sys.path.append(os.path.join(os.path.dirname(__file__), "../../../../"))

from temporeros.models import Temporero
from utils.rut import limpiar_rut, validar_rut


SEPARADOR = "=" * 64
TALLAS_VALIDAS = {"S", "M", "L", "XL"}

VALORES_VERDADEROS = {"Sí", "SI", "si", "sí", "S", "s", "x", "✓", "YES", "yes", "1", 1, True}
VALORES_FALSOS = {"No", "NO", "no", "N", "n", "0", 0, False, "false", "FALSE"}


class Command(BaseCommand):
    help = "Importa temporeros desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)

        # dry-run para simular
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula la importación sin guardar datos"
        )

        # aca si el rut ya existe, esto permite actualizarlo
        parser.add_argument(
            "--update",
            action="store_true",
            help="Actualiza temporeros existentes"
        )

    def parsear_fecha(self, valor):
        if valor is None or str(valor).strip() == "":
            raise ValueError("Fecha vacía")

        if isinstance(valor, datetime):
            return valor.date()

        texto = str(valor).strip()

        formatos = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%d.%m.%Y",
        ]

        for formato in formatos:
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                pass

        raise ValueError(f"Fecha inválida: {valor}")

    def handle(self, *args, **options):
        archivo = options["archivo"]
        dry_run = options["dry_run"]
        update = options["update"]

        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR("Archivo no encontrado"))
            return

        wb = load_workbook(archivo)

        print(SEPARADOR)
        print(f"Importacion de temporeros — {os.path.basename(archivo)}")
        print(SEPARADOR)

        # detecta hoja correcta

        hoja_correcta = None
        nombre_hoja = None

        for nombre in wb.sheetnames:
            ws = wb[nombre]

            for fila in ws.iter_rows(min_row=1, max_row=10):
                valores = [str(c.value).strip() if c.value else "" for c in fila]

                if any("RUT" in v.upper() for v in valores):
                    hoja_correcta = ws
                    nombre_hoja = nombre
                    break

            if hoja_correcta:
                break

        if not hoja_correcta:
            self.stdout.write(self.style.ERROR("No se encontró hoja válida"))
            return

        print(f"Hoja utilizada: {nombre_hoja}")

        # detecta fila de encabezados

        header_row = None

        for idx, fila in enumerate(hoja_correcta.iter_rows(), start=1):
            valores = [str(c.value).strip() if c.value else "" for c in fila]

            if any("RUT" in v.upper() for v in valores):
                header_row = idx
                break

        if not header_row:
            self.stdout.write(self.style.ERROR("No se encontró fila de encabezados"))
            return

        print(f"Fila de encabezados: {header_row}")

        # crea diccionario de encabezados

        headers = {}

        for cell in hoja_correcta[header_row]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column - 1

        total_leidas = 0
        ignoradas = 0
        procesadas = 0
        rut_invalido = 0
        fecha_invalida = 0
        talla_invalida = 0
        creados = 0
        actualizados = 0
        saltados = 0

        # funcion para obtener columnas por nombre

        def obtener(valores, *nombres):
            for nombre in nombres:
                if nombre in headers:
                    idx = headers[nombre]

                    if idx < len(valores):
                        return valores[idx]

            return None

        # booleanos

        def normalizar_bool(valor):
            if valor is None or str(valor).strip() == "":
                return True

            if valor in VALORES_VERDADEROS:
                return True

            if valor in VALORES_FALSOS:
                return False

            return True

        with transaction.atomic():
            # recorre filas de datos

            for fila in hoja_correcta.iter_rows(min_row=header_row + 1):
                total_leidas += 1

                valores = [c.value for c in fila]

                # esto ignora filas vacías
                if all(v is None or str(v).strip() == "" for v in valores):
                    ignoradas += 1
                    continue

                rut = obtener(valores, "RUT", "Rut", "RUT Trabajador")
                nombre = obtener(valores, "Nombre", "Nombre Completo")
                fecha_nacimiento = obtener(valores, "Fecha Nacimiento", "Nacimiento")
                fecha_ingreso = obtener(valores, "Fecha Ingreso", "Ingreso")
                talla = obtener(valores, "Talla", "Talla Polera")
                telefono = obtener(valores, "Teléfono", "Telefono")
                activo = obtener(valores, "Activo", "Estado")
                supervisor = obtener(valores, "Supervisor")

                # aca se hace validacion de RUT

                try:
                    rut_limpio = limpiar_rut(rut)

                    if not validar_rut(rut_limpio):
                        raise ValueError("RUT inválido")

                except ValueError:
                    rut_invalido += 1
                    continue

                # normalización de datos

                nombre_limpio = str(nombre).strip().title() if nombre else ""

                if not nombre_limpio:
                    ignoradas += 1
                    continue

                # talla
                talla_limpia = str(talla).strip().upper() if talla else ""
                if talla_limpia not in TALLAS_VALIDAS:
                    talla_invalida += 1
                    continue

                activo_limpio = normalizar_bool(activo)
                supervisor_limpio = normalizar_bool(supervisor)

                # fechas de nacimiento y ingreso
                try:
                    fecha_nacimiento_limpia = self.parsear_fecha(fecha_nacimiento)
                    fecha_ingreso_limpia = self.parsear_fecha(fecha_ingreso)
                except ValueError:
                    fecha_invalida += 1
                    continue

                telefono_limpio = str(telefono).strip() if telefono else None

                procesadas += 1

                # crear o actualizar

                temporero = Temporero.objects.filter(rut=rut_limpio).first()

                if temporero:
                    if update:
                        temporero.nombre = nombre_limpio
                        temporero.fecha_nacimiento = fecha_nacimiento_limpia
                        temporero.fecha_ingreso = fecha_ingreso_limpia
                        temporero.talla_polera = talla_limpia
                        temporero.telefono = telefono_limpio
                        temporero.activo = activo_limpio
                        temporero.supervisor = supervisor_limpio

                        temporero.save()
                        actualizados += 1
                    else:
                        saltados += 1
                else:
                    Temporero.objects.create(
                        rut=rut_limpio,
                        nombre=nombre_limpio,
                        fecha_nacimiento=fecha_nacimiento_limpia,
                        fecha_ingreso=fecha_ingreso_limpia,
                        talla_polera=talla_limpia,
                        telefono=telefono_limpio,
                        activo=activo_limpio,
                        supervisor=supervisor_limpio
                    )
                    creados += 1

            if dry_run:
                transaction.set_rollback(True)

        print(SEPARADOR)
        print(f"Filas totales leídas:          {total_leidas}")
        print(f"Filas vacías/basura ignoradas: {ignoradas}")
        print(f"Filas procesadas:              {procesadas}")
        print(f"  ├─ Creados:                  {creados}")
        print(f"  ├─ Actualizados:             {actualizados}")
        print(f"  └─ Saltados:                 {saltados}")
        print(f"Filas rechazadas:")
        print(f"  ├─ RUT inválido:             {rut_invalido}")
        print(f"  ├─ Fecha inválida:           {fecha_invalida}")
        print(f"  └─ Talla inválida:           {talla_invalida}")
        print(f"Modo: {'DRY-RUN (no se escribió nada)' if dry_run else 'ESCRITURA EN BD'}")
        print(SEPARADOR)
