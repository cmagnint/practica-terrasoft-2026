import os
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from temporeros.models import Cuartel


SEPARADOR = "=" * 64
VARIEDADES_VALIDAS = {"Duke", "Legacy", "Brigitta", "Star"}

CORRECCIONES_VARIEDAD = {
    "Legasy": "Legacy",
    "Legaci": "Legacy",
    "Bigitta": "Brigitta",
    "Brigita": "Brigitta",
    "Bridgitta": "Brigitta",
    "DUKE": "Duke",
    "duke": "Duke",
    "LEGACY": "Legacy",
    "legacy": "Legacy",
    "BRIGITTA": "Brigitta",
    "brigitta": "Brigitta",
    "STAR": "Star",
    "star": "Star",
}


class Command(BaseCommand):
    help = "Importa cuarteles desde un archivo Excel"

    def add_arguments(self, parser):
        #la ruta del archivo Excel
        parser.add_argument("archivo", type=str)

        #si se ejecuta con --dry-run, simula lo que pasaría con comandos reales, solo que dry run no los guarda, solo hace una simulacion
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula la importación sin guardar datos"
        )

    def handle(self, *args, **options):
        archivo = options["archivo"]

        #avisa si estamos en modo simulación
        dry_run = options["dry_run"]

        #esto comprueba si el archivo existe
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR("Archivo no encontrado"))
            return

        #carga el archivo Excel
        wb = load_workbook(archivo)

        #en una de estas variables se guardará la hoja correcta
        hoja_correcta = None
        nombre_hoja = None

        #iteramos todas las hojas del Excel
        for nombre in wb.sheetnames:
            ws = wb[nombre]

	    #e revisan solo las primeras filas para encontrar encabezados
            for fila in ws.iter_rows(min_row=1, max_row=10):
                valores = [str(c.value).strip() if c.value else "" for c in fila]

       		#busca una fila que tenga "Código" o "Codigo"
                if any("Código" in v or "Codigo" in v for v in valores):
                    hoja_correcta = ws
                    nombre_hoja = nombre
                    break

            # Si se encuentra la hoja, se sale del loop
            if hoja_correcta:
                break

        #si no se encuentra, avisa y termina
        if not hoja_correcta:
            self.stdout.write(self.style.ERROR("No se encontró hoja válida"))
            return

        #muestra qué hoja se utilizará
        print(SEPARADOR)
        print(f"IMPORTACIÓN DE CUARTELES — {os.path.basename(archivo)}")
        print(SEPARADOR)
        print(f"Hoja utilizada: {nombre_hoja}")

        #aqui recorre lasfilas hasta encontrar la que tiene "Código" o "Codigo"
        header_row = None

        for idx, fila in enumerate(hoja_correcta.iter_rows(), start=1):
            valores = [str(c.value).strip() if c.value else "" for c in fila]

            if any("Código" in v or "Codigo" in v for v in valores):
                header_row = idx
                break

        if not header_row:
            self.stdout.write(self.style.ERROR("No se encontró fila de encabezados"))
            return

        #crea diccionario de encabezados
        headers = {}

        #guarda encabezados con índice de columna
        for cell in hoja_correcta[header_row]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column - 1

        print(f"Fila de encabezados: {header_row}")

        #inicializar contadores
        total_leidas = 0
        ignoradas = 0
        procesadas = 0
        creados = 0
        actualizados = 0
        sin_cambios = 0
        duplicados_archivo = 0
        rechazados = 0
        hectareas_invalidas = 0
        codigo_vacio = 0
        variedad_invalida = 0

        codigos_vistos = set()
        cambios_actualizados = []

        #crear carpeta de logs si no existe
        os.makedirs("logs", exist_ok=True)
        fecha_log = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_path = f"logs/import_cuarteles_{fecha_log}.log"

        log = open(log_path, "w", encoding="utf-8")

        #esto lee una columna si existe
        def obtener(valores, *nombres_posibles):
            for nombre_columna in nombres_posibles:
                if nombre_columna in headers:
                    indice = headers[nombre_columna]

                    if indice < len(valores):
                        return valores[indice]

            return None

        #convierte valores tipo Sí/No, 1/0, True/False a boolean
        def normalizar_bool(valor):
            if valor is None or str(valor).strip() == "":
                return True

            verdaderos = {"Sí", "SI", "si", "sí", "S", "s", "x", "✓", "YES", "yes", "1", 1, True}
            falsos = {"No", "NO", "no", "N", "n", "0", 0, False, "false", "FALSE"}

            if valor in verdaderos:
                return True

            if valor in falsos:
                return False

            return True

        #todo el proceso queda dentro de una transacción para poder hacer rollback en dry-run
        with transaction.atomic():
            #empieza a recorrer filas de datos después del header
            for fila in hoja_correcta.iter_rows(min_row=header_row + 1):
                total_leidas += 1

                #esto convierte la fila a lista de valores
                valores = [c.value for c in fila]

                #si hay filas vacías, esto las ignora
                if all(v is None or str(v).strip() == "" for v in valores):
                    ignoradas += 1
                    log.write(f"[FILA {fila[0].row:03}] IGNORADO        : fila vacía\n")
                    continue
		#lee valores desde el excel
                codigo = obtener(valores, "Código Cuartel", "Codigo Cuartel", "Código", "Codigo")
                hectareas = obtener(valores, "Hectáreas", "Hectareas")
                variedad = obtener(valores, "Variedad")
                activo = obtener(valores, "Estado (Activo)", "Activo", "Estado")
		#con esto si no hay codigo, se rechaza
                if codigo is None or str(codigo).strip() == "":
                    ignoradas += 1
                    codigo_vacio += 1
                    log.write(f"[FILA {fila[0].row:03}] RECHAZADO CODIGO: código vacío\n")
                    continue

                codigo = str(codigo).strip().upper()
		#evitamos procesar dos veces el mismo cuartel dentro de excel
                if codigo in codigos_vistos:
                    duplicados_archivo += 1
                    log.write(f"[FILA {fila[0].row:03}] DUPLICADO      : {codigo} ya apareció en el archivo\n")
                    continue

                codigos_vistos.add(codigo)

                try:
                    hectareas_limpias = Decimal(str(hectareas).replace(",", ".").strip())
		    #evita valores invalidos
                    if hectareas_limpias <= 0:
                        raise InvalidOperation

                except (InvalidOperation, AttributeError, ValueError):
                    rechazados += 1
                    hectareas_invalidas += 1
                    log.write(f"[FILA {fila[0].row:03}] RECHAZADO HECT : hectáreas inválidas para {codigo}: {hectareas}\n")
                    continue

                variedad_limpia = str(variedad).strip() if variedad is not None else ""
                variedad_limpia = CORRECCIONES_VARIEDAD.get(variedad_limpia, variedad_limpia.title())

                if variedad_limpia not in VARIEDADES_VALIDAS:
                    rechazados += 1
                    variedad_invalida += 1
                    log.write(f"[FILA {fila[0].row:03}] RECHAZADO VAR  : variedad inválida para {codigo}: {variedad}\n")
                    continue

                activo_limpio = normalizar_bool(activo)

                procesadas += 1

                cuartel = Cuartel.objects.filter(nombre=codigo).first()

                if cuartel:
                    campos_modificados = []

                    if cuartel.hectareas != hectareas_limpias:
                        cuartel.hectareas = hectareas_limpias
                        campos_modificados.append("hectareas")

                    if cuartel.variedad != variedad_limpia:
                        cuartel.variedad = variedad_limpia
                        campos_modificados.append("variedad")

                    if cuartel.activo != activo_limpio:
                        cuartel.activo = activo_limpio
                        campos_modificados.append("activo")

                    if campos_modificados:
                        cuartel.save(update_fields=campos_modificados)
                        actualizados += 1
                        cambios_actualizados.append(f"{codigo}: {'+'.join(campos_modificados)}")
                        log.write(f"[FILA {fila[0].row:03}] ACTUALIZADO    : {codigo} ({', '.join(campos_modificados)})\n")
                    else:
                        sin_cambios += 1
                        log.write(f"[FILA {fila[0].row:03}] SIN CAMBIOS    : {codigo}\n")

                else:
                    Cuartel.objects.create(
                        nombre=codigo,
                        hectareas=hectareas_limpias,
                        variedad=variedad_limpia,
                        activo=activo_limpio
                    )
                    creados += 1
                    log.write(f"[FILA {fila[0].row:03}] OK creado      : {codigo}\n")

            if dry_run:
                transaction.set_rollback(True)

        log.close()

        print(f"Filas leídas:                  {total_leidas}")
        print(f"Filas vacías/basura ignoradas: {ignoradas}")
        print(f"Procesadas:                    {procesadas}")
        print(f"  ├─ Creados:                  {creados}")
        print(f"  ├─ Actualizados:             {actualizados}  ({'; '.join(cambios_actualizados) if cambios_actualizados else 'sin cambios actualizados'})")
        print(f"  ├─ Sin cambios:              {sin_cambios}")
        print(f"  └─ Duplicado en archivo:     {duplicados_archivo}")
        print(f"Rechazados:                    {rechazados}")
        print(f"  ├─ Hectáreas inválidas:      {hectareas_invalidas}")
        print(f"  ├─ Código vacío:             {codigo_vacio}")
        print(f"  └─ Variedad inválida:        {variedad_invalida}")
        print(f"Modo:                          {'DRY-RUN (no se escribió nada)' if dry_run else 'ESCRITURA EN BD'}")
        print(SEPARADOR)
        print(f"Log: {log_path}")
