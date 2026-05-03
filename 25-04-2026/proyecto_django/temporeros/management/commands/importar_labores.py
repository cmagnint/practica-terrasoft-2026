import os
import sys
from datetime import datetime, date

from django.core.management.base import BaseCommand
from django.db import transaction, IntegrityError
from openpyxl import load_workbook

sys.path.append(os.path.join(os.path.dirname(__file__), "../../../../"))

from temporeros.models import Temporero, Cuartel, Labor
from utils.rut import limpiar_rut, validar_rut


SEPARADOR = "=" * 64
TIPOS_VALIDOS = {"Cosecha", "Riego", "Poda", "Pesticida", "Limpieza"}


class Command(BaseCommand):
    help = "Importa labores desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)

        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula sin guardar"
        )

    def parsear_fecha(self, valor):
        if isinstance(valor, datetime):
            return valor.date()

        texto = str(valor).strip()

        formatos = [
            "%d/%m/%Y",
            "%Y-%m-%d",
        ]

        for f in formatos:
            try:
                return datetime.strptime(texto, f).date()
            except:
                pass

        raise ValueError("Fecha inválida")

    def handle(self, *args, **options):
        archivo = options["archivo"]
        dry_run = options["dry_run"]

        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR("Archivo no encontrado"))
            return

        wb = load_workbook(archivo)

        print(SEPARADOR)
        print(f"Importacion de labores — {os.path.basename(archivo)}")
        print(SEPARADOR)

        total = 0
        creadas = 0
        duplicadas = 0
        rechazadas = 0
        sin_temporero = 0
        sin_cuartel = 0

        with transaction.atomic():

            for nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]

                print(f"\nProcesando hoja: {nombre_hoja}")

                header_row = None

                for idx, fila in enumerate(ws.iter_rows(), start=1):
                    valores = [str(c.value).strip() if c.value else "" for c in fila]

                    if any("RUT" in v.upper() for v in valores):
                        header_row = idx
                        break

                if not header_row:
                    print("No tiene encabezado válido, se salta")
                    continue

                headers = {}
                for cell in ws[header_row]:
                    if cell.value:
                        headers[str(cell.value).strip()] = cell.column - 1

                def obtener(valores, *nombres):
                    for n in nombres:
                        if n in headers:
                            i = headers[n]
                            if i < len(valores):
                                return valores[i]
                    return None

                for fila in ws.iter_rows(min_row=header_row + 1):
                    valores = [c.value for c in fila]

                    if all(v is None or str(v).strip() == "" for v in valores):
                        continue

                    total += 1

                    rut = obtener(valores, "RUT")
                    cuartel_nombre = obtener(valores, "Cuartel")
                    tipo = obtener(valores, "Labor", "Tipo")
                    fecha = obtener(valores, "Fecha")
                    horas = obtener(valores, "Horas")
                    kilos = obtener(valores, "Kilos")

                    try:
                        #rut
                        rut = limpiar_rut(rut)
                        if not validar_rut(rut):
                            rechazadas += 1
                            continue

                        temporero = Temporero.objects.filter(rut=rut).first()
                        if not temporero:
                            sin_temporero += 1
                            continue


                        texto = str(cuartel_nombre).upper()
                        cuartel = None

                        for c in Cuartel.objects.all():
                            if c.nombre in texto:
                                cuartel = c
                                break

                        if not cuartel:
                            sin_cuartel += 1
                            continue

                        #tipo
                        tipo = str(tipo).strip().title()
                        if tipo not in TIPOS_VALIDOS:
                            rechazadas += 1
                            continue

                        #fecha
                        fecha = self.parsear_fecha(fecha)
                        if fecha > date.today():
                            rechazadas += 1
                            continue

                        #horas
                        horas = round(float(horas), 2)

                        #kilos
                        if tipo == "Cosecha":
                            kilos = round(float(kilos), 2)
                        else:
                            kilos = None

                    except Exception:
                        rechazadas += 1
                        continue

                    try:
                        labor = Labor(
                            temporero=temporero,
                            cuartel=cuartel,
                            tipo=tipo,
                            fecha=fecha,
                            horas_trabajadas=horas,
                            kilos_cosechados=kilos
                        )

                        labor.full_clean()
                        labor.save()

                        creadas += 1

                    except IntegrityError:
                        duplicadas += 1

                    except Exception:
                        rechazadas += 1

            if dry_run:
                transaction.set_rollback(True)

        print(SEPARADOR)
        print(f"Total filas: {total}")
        print(f"Creadas: {creadas}")
        print(f"Duplicadas: {duplicadas}")
        print(f"Rechazadas: {rechazadas}")
        print(f"Sin temporero: {sin_temporero}")
        print(f"Sin cuartel: {sin_cuartel}")
        print(f"Modo: {'DRY-RUN' if dry_run else 'REAL'}")
        print(SEPARADOR)
