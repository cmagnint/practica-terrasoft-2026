import os
import sys
from datetime import datetime
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

sys.path.append(os.path.join(os.path.dirname(__file__), "../../../../"))

from temporeros.models import Temporero
from utils.rut import limpiar_rut, validar_rut


SEPARADOR = "=" * 64
TALLAS_VALIDAS = {"S", "M", "L", "XL"}

VALORES_VERDADEROS = {"Sí", "SI", "si", "sí", "S", "s", "x", "✓", "YES", "yes", "1", 1, True}
VALORES_FALSOS = {"No", "NO", "no", "N", "n", "0", 0, False, "false", "FALSE"}


class Command(BaseCommand):
    help = "Importa temporeros desde un archivo Excel"

    def add_arguments(self, parser):
        #ruta del archivo excel
        parser.add_argument("archivo", type=str)

        #simula la importacion sin guardar en bd
        parser.add_argument("--dry-run", action="store_true")

        #permite actualizar temporeros existentes
        parser.add_argument("--update", action="store_true")

    def parsear_fecha(self, valor):
        #convierte fechas en distintos formatos a date
        if valor is None or str(valor).strip() == "":
            raise ValueError("Fecha vacía")

        if isinstance(valor, datetime):
            return valor.date()

        texto = str(valor).strip()

        formatos = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y"]

        for formato in formatos:
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                pass

        raise ValueError(f"Fecha inválida: {valor}")

    def normalizar_bool(self, valor, default=False):
        #convierte valores del excel a booleanos y rechaza valores desconocidos
        if valor is None or str(valor).strip() == "":
            return default

        if valor in VALORES_VERDADEROS:
            return True

        if valor in VALORES_FALSOS:
            return False

        texto = str(valor).strip()

        if texto in VALORES_VERDADEROS:
            return True

        if texto in VALORES_FALSOS:
            return False

        raise ValueError(f"valor no reconocido: {valor}")

    def handle(self, *args, **options):
        archivo = options["archivo"]
        dry_run = options["dry_run"]
        update = options["update"]

        #verifica que el archivo exista
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR("Archivo no encontrado"))
            return

        wb = load_workbook(archivo)

        #crea carpeta de logs
        os.makedirs("logs", exist_ok=True)

        #crea archivo log con fecha y hora
        log_path = f"logs/import_temporeros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log = open(log_path, "w", encoding="utf-8")

        print(SEPARADOR)
        print(f"IMPORTACIÓN DE TEMPOREROS — {os.path.basename(archivo)}")
        print(SEPARADOR)

        #busca la hoja correcta detectando una fila con rut
        hoja = None
        nombre_hoja = None

        for nombre in wb.sheetnames:
            ws = wb[nombre]

            for fila in ws.iter_rows(min_row=1, max_row=10):
                valores = [str(c.value).strip() if c.value else "" for c in fila]

                if any("RUT" in v.upper() for v in valores):
                    hoja = ws
                    nombre_hoja = nombre
                    break

            if hoja:
                break

        if not hoja:
            print("No se encontró hoja válida")
            log.close()
            return

        #busca la fila real de encabezados
        header_row = None

        for idx, fila in enumerate(hoja.iter_rows(), start=1):
            valores = [str(c.value).strip() if c.value else "" for c in fila]

            if any("RUT" in v.upper() for v in valores):
                header_row = idx
                break

        if not header_row:
            print("No se encontró fila de encabezados")
            log.close()
            return

        #crea diccionario columna -> indice
        headers = {}

        for cell in hoja[header_row]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column - 1

        #obtiene un valor aceptando varios nombres posibles de columna
        def obtener(valores, *nombres):
            for nombre in nombres:
                if nombre in headers:
                    indice = headers[nombre]

                    if indice < len(valores):
                        return valores[indice]

            return None

        total_leidas = 0
        ignoradas = 0
        procesadas = 0
        creados = 0
        actualizados = 0
        saltados = 0
        rechazados = 0

        rut_invalidos = 0
        fechas_invalidas = 0
        tallas_invalidas = 0
        booleanos_invalidos = 0

        with transaction.atomic():
            #recorre filas desde despues del encabezado
            for fila in hoja.iter_rows(min_row=header_row + 1):
                total_leidas += 1

                valores = [c.value for c in fila]
                fila_excel = fila[0].row

                #ignora filas vacias o con solo espacios
                if all(v is None or str(v).strip() == "" for v in valores):
                    ignoradas += 1
                    log.write(f"[FILA {fila_excel:03}] IGNORADO        : fila vacía\n")
                    continue

                nombre_raw = obtener(valores, "Nombre", "Nombre Completo", "Trabajador")
                nombre = str(nombre_raw).strip().title() if nombre_raw else ""

                if not nombre or nombre.upper().startswith(("TOTAL", "VERSIÓN", "APROBADO")):
                    ignoradas += 1
                    log.write(f"[FILA {fila_excel:03}] IGNORADO        : fila basura\n")
                    continue

                try:
                    # limpia y valida rut
                    rut_raw = obtener(valores, "RUT", "Rut", "RUT Trabajador")
                    rut = limpiar_rut(rut_raw)

                    if not validar_rut(rut):
                        rut_invalidos += 1
                        raise ValueError(f"RUT inválido: {rut_raw}")

                    #parsea fechas
                    try:
                        fecha_nacimiento = self.parsear_fecha(
                            obtener(valores, "Fecha Nacimiento", "Nacimiento", "Fecha de Nacimiento")
                        )
                        fecha_ingreso = self.parsear_fecha(
                            obtener(valores, "Fecha Ingreso", "Ingreso", "Fecha de Ingreso")
                        )
                    except ValueError as e:
                        fechas_invalidas += 1
                        raise ValueError(e)

                    #valida las tallas
                    talla = obtener(valores, "Talla", "Talla Polera", "Talla polera")
                    talla = str(talla).strip().upper() if talla else ""

                    if talla not in TALLAS_VALIDAS:
                        tallas_invalidas += 1
                        raise ValueError(f"Talla inválida: {talla}")

                    #limpia telefono
                    telefono = obtener(valores, "Teléfono", "Telefono")
                    telefono = str(telefono).replace("+", "").replace(" ", "").replace("(", "").replace(")", "").strip() if telefono else None
                    telefono = telefono if telefono else None

                    #limpia contacto de emergencia
                    contacto = obtener(valores, "Contacto Emergencia", "Contacto emergencia", "Contacto")
                    contacto = str(contacto).strip() if contacto else None
                    contacto = contacto if contacto else None

                    #normaliza activo y supervisor
                    try:
                        activo = self.normalizar_bool(obtener(valores, "Activo", "Estado"), default=True)
                        supervisor = self.normalizar_bool(obtener(valores, "Supervisor"), default=False)
                    except ValueError as e:
                        booleanos_invalidos += 1
                        raise ValueError(e)

                except ValueError as e:
                    rechazados += 1
                    log.write(f"[FILA {fila_excel:03}] RECHAZADO       : {e} → {nombre}\n")
                    continue

                procesadas += 1

                existente = Temporero.objects.filter(rut=rut).first()

                if existente:
                    if update:
                        #actualiza registro existente
                        existente.nombre = nombre
                        existente.fecha_nacimiento = fecha_nacimiento
                        existente.fecha_ingreso = fecha_ingreso
                        existente.talla_polera = talla
                        existente.telefono = telefono
                        existente.contacto_emergencia = contacto
                        existente.activo = activo
                        existente.supervisor = supervisor
                        existente.save()

                        actualizados += 1
                        log.write(f"[FILA {fila_excel:03}] ACTUALIZADO     : {nombre} ({rut})\n")
                    else:
                        #salta existentes si no se usa --update
                        saltados += 1
                        log.write(f"[FILA {fila_excel:03}] SALTADO         : ya existe {rut} → {nombre}\n")

                else:
                    #crea nuevo temporero
                    Temporero.objects.create(
                        rut=rut,
                        nombre=nombre,
                        fecha_nacimiento=fecha_nacimiento,
                        fecha_ingreso=fecha_ingreso,
                        talla_polera=talla,
                        telefono=telefono,
                        contacto_emergencia=contacto,
                        activo=activo,
                        supervisor=supervisor,
                    )

                    creados += 1
                    log.write(f"[FILA {fila_excel:03}] OK creado       : {nombre} ({rut})\n")

            #si es un dryrun se revierten los cambios
            if dry_run:
                transaction.set_rollback(True)

        log.close()

        print(f"Hoja utilizada:               {nombre_hoja}")
        print(f"Fila de encabezados:          {header_row}")
        print(f"Filas totales leídas:         {total_leidas}")
        print(f"Filas vacías/basura ignoradas:{ignoradas}")
        print(f"Filas procesadas:             {procesadas}")
        print(f"  ├─ Creados:                 {creados}")
        print(f"  ├─ Actualizados:            {actualizados}")
        print(f"  └─ Saltados:                {saltados}")
        print(f"Filas rechazadas:             {rechazados}")
        print(f"  ├─ RUT inválido:            {rut_invalidos}")
        print(f"  ├─ Fecha inválida:          {fechas_invalidas}")
        print(f"  ├─ Talla inválida:          {tallas_invalidas}")
        print(f"  └─ Booleano inválido:       {booleanos_invalidos}")
        print(f"Modo:                         {'DRY-RUN (no se escribió nada)' if dry_run else 'ESCRITURA EN BD'}")
        print(SEPARADOR)
        print(f"Log: {log_path}")
