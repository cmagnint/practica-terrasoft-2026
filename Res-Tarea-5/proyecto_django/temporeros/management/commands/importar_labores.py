import os
import sys
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction, IntegrityError
from django.core.exceptions import ValidationError
from openpyxl import load_workbook

#permite importar utils/rut.py desde la carpeta 25-04-2026
sys.path.append(os.path.join(os.path.dirname(__file__), "../../../../"))

from temporeros.models import Temporero, Cuartel, Labor
from utils.rut import limpiar_rut, validar_rut


SEPARADOR = "=" * 64
TIPOS_VALIDOS = {"Cosecha", "Poda", "Riego", "Pesticida", "Limpieza"}


class Command(BaseCommand):
    help = "Importa labores desde un archivo Excel"

    def add_arguments(self, parser):
        #ruta del archivo excel
        parser.add_argument("archivo", type=str)

        #simula sin guardar cambios
        parser.add_argument("--dry-run", action="store_true")

        #procesa una sola hoja si se indica
        parser.add_argument("--hoja", type=str, default=None)

    def normalizar_texto(self, texto):
        #normaliza encabezados para comparar sin depender de mayusculas, tildes o simbolos
        texto = str(texto).strip().lower()
        texto = texto.replace("á", "a").replace("é", "e").replace("í", "i")
        texto = texto.replace("ó", "o").replace("ú", "u").replace("°", "")
        texto = texto.replace(".", "")
        texto = texto.replace("(", "").replace(")", "")
        return texto

    def fecha_desde_nombre_hoja(self, nombre_hoja):
        #extrae fecha desde nombres como "Lunes 30-03"
        match = re.search(r"(\d{2})-(\d{2})", nombre_hoja)

        if not match:
            raise ValueError(f"No se pudo extraer fecha de la hoja: {nombre_hoja}")

        dia = int(match.group(1))
        mes = int(match.group(2))

        return date(2026, mes, dia)

    def parsear_fecha(self, valor):
        #convierte fechas en distintos formatos a date
        if valor is None or str(valor).strip() == "":
            raise ValueError("Fecha vacía")

        if isinstance(valor, datetime):
            return valor.date()

        texto = str(valor).strip()

        formatos = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%d.%m.%Y",
        ]

        for formato in formatos:
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                pass

        raise ValueError(f"Fecha inválida: {valor}")

    def parsear_numero(self, valor):
        #convierte valores como "7,5", "8 hrs" o "120 kg" a decimal
        if valor is None or str(valor).strip() == "":
            raise ValueError("Número vacío")

        texto = str(valor).strip().replace(",", ".")

        #deja solo numeros, punto decimal y signo negativo
        texto = re.sub(r"[^0-9.\-]", "", texto)

        if texto in {"", "-", ".", "-."}:
            raise ValueError(f"No es número: {valor}")

        try:
            return Decimal(texto)
        except InvalidOperation:
            raise ValueError(f"Número inválido: {valor}")

    def normalizar_cuartel(self, valor):
        #normaliza cuarteles: "Cuartel A-1", "SECTOR A2" o "A1" -> "A-1"
        if valor is None or str(valor).strip() == "":
            raise ValueError("Cuartel vacío")

        texto = str(valor).strip()

        prefijos = [
            "CUARTEL ",
            "Cuartel ",
            "cuartel ",
            "SECTOR ",
            "Sector ",
            "sector ",
        ]

        for prefijo in prefijos:
            if texto.startswith(prefijo):
                texto = texto[len(prefijo):]

        texto = texto.strip().upper().replace(" ", "")

        if re.match(r"^[A-Z]\d+$", texto):
            texto = f"{texto[0]}-{texto[1:]}"

        return texto

    def normalizar_tipo(self, valor):
        #normaliza tipo de labor y traduce valores conocidos
        if valor is None or str(valor).strip() == "":
            raise ValueError("Tipo vacío")

        texto = str(valor).strip()

        traducciones = {
            "harvesting": "Cosecha",
            "harvest": "Cosecha",
        }

        texto_lower = texto.lower()

        if texto_lower in traducciones:
            return traducciones[texto_lower]

        tipo = texto.capitalize()

        if tipo not in TIPOS_VALIDOS:
            raise ValueError(f"Tipo inválido: {valor}")

        return tipo

    def mapear_columnas(self, header_row):
        #asocia columnas del excel con campos internos usando variantes flexibles
        resultado = {}

        variantes_normalizadas = {
            "rut": ["rut", "n rut", "rut trabajador"],
            "nombre": ["trabajador", "nombre", "nombre completo", "nombre del trabajador"],
            "cuartel": ["cuartel", "sector", "cuartel/sector"],
            "tipo": ["tipo labor", "tipo de labor", "tipo", "tipo de tarea", "labor", "tarea"],
            "horas": ["horas", "hrs", "horas trabajadas", "horas trab"],
            "kilos": ["kilos", "kg", "kg cosechados", "kilos cosechados", "kilos solo cosecha"],
            "fecha": ["fecha", "fecha de trabajo"],
            "obs": ["observaciones", "obs", "notas"],
        }

        for cell in header_row:
            if not cell.value:
                continue

            nombre_col = self.normalizar_texto(cell.value)

            for key, variantes in variantes_normalizadas.items():
                if any(variante in nombre_col for variante in variantes):
                    resultado[key] = cell.column - 1
                    break

        return resultado

    def handle(self, *args, **options):
        archivo = options["archivo"]
        dry_run = options["dry_run"]
        hoja_unica = options["hoja"]

        #valida existencia del archivo
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR("Archivo no encontrado"))
            return

        wb = load_workbook(archivo)

        #crea carpeta de logs
        os.makedirs("logs", exist_ok=True)

        #crea log con fecha y hora
        log_path = f"logs/import_labores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log = open(log_path, "w", encoding="utf-8")

        print(SEPARADOR)
        print(f"IMPORTACIÓN DE LABORES — {os.path.basename(archivo)}")
        print(SEPARADOR)

        total_leidas = 0
        total_importadas = 0
        total_rechazadas = 0
        total_duplicadas = 0
        total_vacias = 0

        #define si se procesan todas las hojas o solo una
        if hoja_unica:
            if hoja_unica not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f"No existe la hoja: {hoja_unica}"))
                log.close()
                return

            hojas_a_procesar = [hoja_unica]
        else:
            hojas_a_procesar = wb.sheetnames

        with transaction.atomic():
            for nombre_hoja in hojas_a_procesar:
                ws = wb[nombre_hoja]

                #separador por hoja en log
                log.write(f"\n--- Hoja: {nombre_hoja} ---\n")

                leidas = 0
                importadas = 0
                rechazadas = 0
                duplicadas = 0
                vacias = 0

                #intenta obtener fecha desde el nombre de la hoja
                try:
                    fecha_hoja = self.fecha_desde_nombre_hoja(nombre_hoja)
                except ValueError:
                    fecha_hoja = None

                #detecta fila de encabezados
                header_row = None

                for idx, fila in enumerate(ws.iter_rows(), start=1):
                    valores = [self.normalizar_texto(c.value) if c.value else "" for c in fila]

                    if any(v in ["rut", "n rut", "trabajador", "nombre", "nombre completo"] for v in valores):
                        header_row = idx
                        break

                if not header_row:
                    print(f"\nHoja: {nombre_hoja} — sin encabezado válido, se omite")
                    log.write(f"[HOJA {nombre_hoja}] OMITIDA: sin encabezado válido\n")
                    continue

                #mapea columnas de forma flexible
                columnas = self.mapear_columnas(ws[header_row])
                tiene_col_fecha = "fecha" in columnas

                print(f"\nHoja: {nombre_hoja}")

                #recorre filas de datos
                for fila in ws.iter_rows(min_row=header_row + 1):
                    leidas += 1
                    total_leidas += 1

                    valores = [c.value for c in fila]
                    fila_excel = fila[0].row

                    #ignora filas vacias
                    if all(v is None or str(v).strip() == "" for v in valores):
                        vacias += 1
                        total_vacias += 1
                        log.write(f"[FILA {fila_excel:03}] IGNORADO        : fila vacía\n")
                        continue

                    #obtiene valor segun campo ya mapeado
                    def get(campo):
                        if campo not in columnas:
                            return None

                        indice = columnas[campo]

                        if indice >= len(valores):
                            return None

                        return valores[indice]

                    try:
                        #valida rut
                        rut = limpiar_rut(get("rut"))

                        if not validar_rut(rut):
                            raise ValueError(f"RUT inválido: {rut}")

                        #valida existencia de temporero
                        temporero = Temporero.objects.filter(rut=rut).first()

                        if not temporero:
                            raise LookupError(f"Temporero no existe: {rut}")

                        #valida existencia de cuartel
                        nombre_cuartel = self.normalizar_cuartel(get("cuartel"))
                        cuartel = Cuartel.objects.filter(nombre=nombre_cuartel).first()

                        if not cuartel:
                            raise LookupError(f"Cuartel no existe: {nombre_cuartel}")

                        #normaliza tipo
                        tipo = self.normalizar_tipo(get("tipo"))

                        #obtiene fecha desde columna o desde nombre de hoja
                        valor_fecha = get("fecha")

                        if tiene_col_fecha and valor_fecha not in [None, ""]:
                            fecha = self.parsear_fecha(valor_fecha)
                        else:
                            if not fecha_hoja:
                                raise ValueError("No hay fecha en columna ni en nombre de hoja")

                            fecha = fecha_hoja

                        #valida fechas de negocio
                        if fecha > date.today():
                            raise ValueError(f"Fecha futura: {fecha}")

                        if fecha < temporero.fecha_ingreso:
                            raise ValueError(f"Fecha anterior al ingreso: {fecha}")

                        #valida horas
                        horas = self.parsear_numero(get("horas"))

                        if horas <= 0 or horas > 12:
                            raise ValueError(f"Horas fuera de rango: {horas}")

                        #valida kilos segun tipo
                        valor_kilos = get("kilos")

                        if tipo == "Cosecha":
                            kilos = self.parsear_numero(valor_kilos)

                            if kilos < 0:
                                raise ValueError(f"Kilos negativos: {kilos}")
                        else:
                            kilos = None

                        #observaciones opcionales
                        obs = get("obs")
                        observaciones = str(obs).strip() if obs else ""

                    except ValueError as e:
                        rechazadas += 1
                        total_rechazadas += 1
                        log.write(f"[FILA {fila_excel:03}] RECHAZADO       : {e}\n")
                        continue

                    except LookupError as e:
                        rechazadas += 1
                        total_rechazadas += 1
                        log.write(f"[FILA {fila_excel:03}] RECHAZADO FK    : {e}\n")
                        continue

                    try:
                        #crea labor y aplica validaciones del modelo
                        labor = Labor(
                            temporero=temporero,
                            cuartel=cuartel,
                            tipo=tipo,
                            fecha=fecha,
                            horas_trabajadas=horas,
                            kilos_cosechados=kilos,
                            observaciones=observaciones,
                        )

                        labor.full_clean()
                        labor.save()

                        importadas += 1
                        total_importadas += 1
                        log.write(
                            f"[FILA {fila_excel:03}] OK creado       : "
                            f"{rut} {nombre_cuartel} {tipo} {fecha}\n"
                        )

                    except IntegrityError:
                        duplicadas += 1
                        total_duplicadas += 1
                        log.write(f"[FILA {fila_excel:03}] DUPLICADO      : UniqueConstraint\n")

                    except ValidationError as e:
                        rechazadas += 1
                        total_rechazadas += 1
                        log.write(f"[FILA {fila_excel:03}] RECHAZADO VAL  : {e}\n")

                print(
                    f"  Leídas: {leidas} | Importadas: {importadas} | "
                    f"Rechazadas: {rechazadas} | Duplicadas: {duplicadas} | Vacías: {vacias}"
                )

            #revierte cambios si es dry-run
            if dry_run:
                transaction.set_rollback(True)

        log.close()

        print(SEPARADOR)
        print("TOTALES")
        print(f"  Filas leídas:               {total_leidas}")
        print(f"  Importadas:                 {total_importadas}")
        print(f"  Rechazadas:                 {total_rechazadas}")
        print(f"  Duplicadas:                 {total_duplicadas}")
        print(f"  Vacías:                     {total_vacias}")
        print(f"Modo: {'DRY-RUN (no se escribió nada)' if dry_run else 'ESCRITURA EN BD'}")
        print(SEPARADOR)
        print(f"Log: {log_path}")
